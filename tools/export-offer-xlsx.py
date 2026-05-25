from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from statistics import mean

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
PRODUCTS_PATH = ROOT / "assets" / "data" / "products.json"
OUT_DIR = ROOT / "oferty"
OUT_PATH = OUT_DIR / "FAPO_Polska_katalog_kalkulacja_2026-05-25.xlsx"


COLORS = {
    "brand_dark": "12171D",
    "light_fill": "F5F7FA",
    "light_line": "D9E0E8",
    "orange": "F54E1A",
    "orange_2": "FF7A2B",
    "white": "FFFFFF",
}


def load_products() -> list[dict]:
    products = json.loads(PRODUCTS_PATH.read_text(encoding="utf-8"))
    return sorted(
        products,
        key=lambda product: (
            product.get("category") or "",
            product.get("sku") or "",
            product.get("title") or "",
        ),
    )


def fill(cell, color: str) -> None:
    cell.fill = PatternFill("solid", fgColor=color)


def style_header(cell) -> None:
    fill(cell, COLORS["brand_dark"])
    cell.font = Font(name="Arial", bold=True, color=COLORS["white"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color=COLORS["orange"]))


def style_body(cell, row_index: int) -> None:
    fill(cell, COLORS["white"] if row_index % 2 else COLORS["light_fill"])
    cell.border = Border(bottom=Side(style="thin", color=COLORS["light_line"]))
    cell.alignment = Alignment(vertical="top")


def build_offer_sheet(wb: Workbook, products: list[dict]) -> None:
    ws = wb.active
    ws.title = "Oferta"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"

    ws.merge_cells("A1:H1")
    ws["A1"] = "FAPO POLSKA"
    ws["A1"].font = Font(name="Arial", size=28, bold=True, color=COLORS["white"])
    fill(ws["A1"], COLORS["brand_dark"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:H2")
    ws["A2"] = "Katalog części i kalkulacja oferty"
    ws["A2"].font = Font(name="Arial", size=18, bold=True, color=COLORS["orange_2"])
    fill(ws["A2"], COLORS["brand_dark"])
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 30

    summary_rows = [
        ("Data przygotowania", date(2026, 5, 25).isoformat()),
        ("Liczba produktów w katalogu", len(products)),
        ("Waluta", "PLN"),
        ("Źródło danych", "assets/data/products.json"),
        ("Kontakt handlowy", "office@fapomoto.pl | info@fapomoto.pl"),
        (
            "Instrukcja",
            "W arkuszu Katalog i kalkulacja wpisz rabat oraz ilość. "
            "Excel policzy cenę po rabacie i wartość pozycji.",
        ),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=4):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, value)
        ws.cell(row_index, 1).font = Font(bold=True, color=COLORS["brand_dark"])
        ws.cell(row_index, 2).font = Font(color="1F2937")
        fill(ws.cell(row_index, 1), "FFF0E8")
        fill(ws.cell(row_index, 2), COLORS["light_fill"])
        for col in range(1, 9):
            cell = ws.cell(row_index, col)
            cell.border = Border(bottom=Side(style="thin", color=COLORS["light_line"]))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.merge_cells("B9:C9")
    ws["A9"] = "Kalkulacja klienta"
    ws["B9"] = "Wartość"
    for cell in ws[9]:
        fill(cell, COLORS["brand_dark"])
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.alignment = Alignment(horizontal="center")

    last_row = len(products) + 1
    calc_rows = [
        ("Wybrane pozycje", f'=COUNTIF(\'Katalog i kalkulacja\'!$I$2:$I${last_row},">0")', "#,##0"),
        ("Łączna ilość", f"=SUM('Katalog i kalkulacja'!$I$2:$I${last_row})", "#,##0"),
        (
            "Suma katalogowa PLN",
            f"=SUMPRODUCT('Katalog i kalkulacja'!$F$2:$F${last_row},'Katalog i kalkulacja'!$I$2:$I${last_row})",
            "#,##0.00 zł",
        ),
        ("Suma po rabatach PLN", f"=SUM('Katalog i kalkulacja'!$J$2:$J${last_row})", "#,##0.00 zł"),
    ]
    for row_index, (label, formula, number_format) in enumerate(calc_rows, start=10):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, formula)
        ws.cell(row_index, 1).font = Font(bold=True, color=COLORS["brand_dark"])
        ws.cell(row_index, 2).font = Font(bold=True, color=COLORS["orange"])
        ws.cell(row_index, 2).number_format = number_format
        for col in range(1, 4):
            fill(ws.cell(row_index, col), COLORS["light_fill"])
            ws.cell(row_index, col).border = Border(bottom=Side(style="thin", color=COLORS["light_line"]))

    category_prices: dict[str, list[float]] = {}
    for product in products:
        category = product.get("category") or "Brak kategorii"
        category_prices.setdefault(category, []).append(float(product.get("priceFrom") or 0))

    headers = ["Kategoria", "Liczba produktów", "Cena min PLN", "Cena średnia PLN", "Cena max PLN", "Wartość wybrana PLN"]
    for col, header in enumerate(headers, 1):
        style_header(ws.cell(16, col, header))

    for row_index, category in enumerate(sorted(category_prices), start=17):
        prices = category_prices[category]
        values = [
            category,
            len(prices),
            min(prices),
            round(mean(prices), 2),
            max(prices),
            f"=SUMIF('Katalog i kalkulacja'!$D:$D,A{row_index},'Katalog i kalkulacja'!$J:$J)",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_index, col, value)
            style_body(cell, row_index)
            if col >= 3:
                cell.number_format = "#,##0.00 zł"
            if col == 2:
                cell.number_format = "#,##0"

    for col, width in enumerate([24, 26, 18, 18, 18, 20, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_catalog_sheet(wb: Workbook, products: list[dict]) -> None:
    ws = wb.create_sheet("Katalog i kalkulacja")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "LP",
        "ID",
        "SKU",
        "Kategoria",
        "Nazwa produktu",
        "Cena katalogowa PLN",
        "Rabat %",
        "Cena po rabacie PLN",
        "Ilość",
        "Wartość PLN",
        "Link produktu FAPO PL",
        "Link zdjęcia",
        "Liczba zdjęć",
        "Źródło",
        "Uwagi",
    ]
    for col, header in enumerate(headers, 1):
        style_header(ws.cell(1, col, header))
    ws.row_dimensions[1].height = 36

    for row_index, product in enumerate(products, start=2):
        price = float(product.get("priceFrom") or 0)
        product_url = product.get("canonicalUrl") or f"https://fapomoto.pl/{(product.get('url') or '').lstrip('/')}"
        image_url = product.get("image") or (product.get("images") or [""])[0]
        images_count = len(product.get("images") or [])
        row_values = [
            row_index - 1,
            product.get("id") or "",
            (product.get("sku") or "").strip(),
            (product.get("category") or "").strip(),
            (product.get("title") or "").strip(),
            price,
            0,
            f"=ROUND(F{row_index}*(1-G{row_index}),2)",
            0,
            f"=ROUND(H{row_index}*I{row_index},2)",
            product_url,
            image_url,
            images_count,
            product.get("source") or "",
            "",
        ]
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row_index, col, value)
            style_body(cell, row_index)
            if col in (5, 11, 12, 15):
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.cell(row_index, 6).number_format = "#,##0.00 zł"
        ws.cell(row_index, 7).number_format = "0.00%"
        ws.cell(row_index, 8).number_format = "#,##0.00 zł"
        ws.cell(row_index, 9).number_format = "#,##0"
        ws.cell(row_index, 10).number_format = "#,##0.00 zł"

        if product_url:
            ws.cell(row_index, 11).hyperlink = product_url
            ws.cell(row_index, 11).style = "Hyperlink"
        if image_url:
            ws.cell(row_index, 12).hyperlink = image_url
            ws.cell(row_index, 12).style = "Hyperlink"

    last_row = len(products) + 1
    table_ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    table = Table(displayName="TabelaKatalogFAPO", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = table_ref

    ws.conditional_formatting.add(
        f"I2:I{last_row}",
        FormulaRule(formula=["I2>0"], fill=PatternFill("solid", fgColor="FFF0E8")),
    )
    ws.conditional_formatting.add(
        f"J2:J{last_row}",
        FormulaRule(formula=["J2>0"], fill=PatternFill("solid", fgColor="FFE0D1")),
    )

    widths = {
        1: 8,
        2: 10,
        3: 16,
        4: 18,
        5: 56,
        6: 18,
        7: 12,
        8: 20,
        9: 10,
        10: 16,
        11: 58,
        12: 52,
        13: 12,
        14: 14,
        15: 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_terms_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Warunki")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "FAPO Polska - warunki do oferty"
    ws["A1"].font = Font(size=20, bold=True, color=COLORS["white"])
    fill(ws["A1"], COLORS["brand_dark"])
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    terms = [
        ("Ceny", "Ceny w PLN według aktualnego katalogu FAPO Polska. Arkusz nie rozdziela automatycznie wartości netto/brutto."),
        ("Kalkulacja", "Wpisz ilość i ewentualny rabat przy wybranych pozycjach. Wartość pozycji i suma oferty liczą się automatycznie."),
        ("Gwarancja", "Standardowa gwarancja: 12 miesięcy. Dla kwalifikujących się produktów możliwe rozszerzenie do 24 miesięcy."),
        ("Zwroty", "Zwrot wymaga wcześniejszej autoryzacji RA. Standardowo do 30 dni od doręczenia, zgodnie z polityką FAPO Polska."),
        ("Realizacja", "Przetwarzanie zamówień zwykle 1-2 dni robocze. Dostawa po nadaniu zwykle 2-7 dni roboczych."),
        ("Kontakt", "office@fapomoto.pl | info@fapomoto.pl"),
        ("Dane firmy", "FAPO Polska | ul. Bokserska 59 lok. 17, 02-690 Warszawa | NIP 9512184841 | REGON 141812444"),
    ]
    for row_index, (label, text) in enumerate(terms, start=3):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, text)
        ws.cell(row_index, 1).font = Font(bold=True, color=COLORS["brand_dark"])
        fill(ws.cell(row_index, 1), "FFF0E8")
        fill(ws.cell(row_index, 2), COLORS["light_fill"])
        ws.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 7):
            ws.cell(row_index, col).border = Border(bottom=Side(style="thin", color=COLORS["light_line"]))

    for col, width in enumerate([20, 95, 15, 15, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    products = load_products()

    wb = Workbook()
    wb.properties.creator = "FAPO Polska"
    wb.properties.title = "FAPO Polska - katalog i kalkulacja części"
    wb.properties.subject = "Katalog wszystkich części z kalkulacją oferty"
    wb.properties.keywords = "FAPO, katalog, kalkulacja, automotive, performance parts"

    build_offer_sheet(wb, products)
    build_catalog_sheet(wb, products)
    build_terms_sheet(wb)

    ws_catalog = wb["Katalog i kalkulacja"]
    assert ws_catalog.max_row == len(products) + 1
    assert ws_catalog.max_column == 15

    wb.save(OUT_PATH)

    check = load_workbook(OUT_PATH, data_only=False, read_only=True)
    assert check["Katalog i kalkulacja"].max_row == len(products) + 1
    assert check["Katalog i kalkulacja"].max_column == 15
    check.close()

    print(OUT_PATH)
    print(f"products={len(products)}")
    print(f"size_bytes={OUT_PATH.stat().st_size}")


if __name__ == "__main__":
    main()
